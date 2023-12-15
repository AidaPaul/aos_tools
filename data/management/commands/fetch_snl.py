import json
from copy import copy

import requests
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from data.models import (
    Event,
    BCP,
    W40K,
    BOLT_ACTION,
    AOS,
    Pairing,
    List,
    Participant,
    Player,
    SNL,
)

import openpyxl


# Function to parse the snl_results.xlsx file and build a Python dictionary
def parse_snl_results_file(filename):
    # Load the Excel file
    wb = openpyxl.load_workbook(filename)

    # Get the workbook sheets
    tournaments_sheet = wb["Tournaments"]
    players_sheet = wb["Players"]
    pairings_sheet = wb["Pairings"]

    # Iterate through the tournaments sheet and build a list of tournaments
    tournaments = []
    for row in tournaments_sheet.iter_rows(min_row=2):
        tournament_id = str(row[0].value)
        name = row[1].value
        attendee_count = int(row[2].value)
        start_date = row[3].value
        country = row[4].value
        state = row[5].value
        location = row[6].value
        max_players = int(row[7].value)

        tournament = {
            "tournamentID": tournament_id,
            "name": name,
            "attendeeCount": attendee_count,
            "startDate": start_date,
            "country": country,
            "state": state,
            "location": location,
            "maxPlayers": max_players,
        }

        tournaments.append(tournament)

    # Iterate through the players sheet and build a dictionary of players
    players = {}
    for row in players_sheet.iter_rows(min_row=2):
        tournament_id = str(row[0].value)
        registration_id = str(row[1].value)
        player_name = row[2].value
        faction = row[3].value
        army = row[4].value
        club = row[5].value
        list_plain_text = row[6].value

        player = {
            "tournamentID": tournament_id,
            "registrationID": registration_id,
            "playerName": player_name,
            "faction": faction,
            "army": army,
            "club": club,
            "listPlainText": list_plain_text,
        }

        if tournament_id in players:
            players[tournament_id].append(player)
        else:
            players[tournament_id] = [player]

    # Iterate through the pairings sheet and build a dictionary of pairings
    pairings = {}
    for row in pairings_sheet.iter_rows(min_row=2):
        tournament_id = str(row[0].value)
        round_number = int(row[1].value)
        player_id1 = str(row[2].value)
        player_result1 = row[3].value
        player_score1 = int(row[4].value)
        player_bonus1 = int(row[5].value)
        player_id2 = str(row[6].value)
        player_result2 = row[7].value
        player_score2 = int(row[8].value)
        player_bonus2 = int(row[9].value)

        pairing = {
            "tournamentID": tournament_id,
            "roundNumber": round_number,
            "playerId1": player_id1,
            "playerResult1": player_result1,
            "playerScore1": player_score1,
            "playerBonus1": player_bonus1,
            "playerId2": player_id2,
            "playerResult2": player_result2,
            "playerScore2": player_score2,
            "playerBonus2": player_bonus2,
        }

        if tournament_id in pairings:
            pairings[tournament_id].append(pairing)
        else:
            pairings[tournament_id] = [pairing]

    # Return the parsed data
    return tournaments, players, pairings


def store_tournaments(tournaments):
    for tournament in tournaments:
        source = SNL
        source_id = tournament["tournamentID"]
        source_json = tournament

        event, _ = Event.objects.update_or_create(
            source=source,
            source_id=source_id,
            source_json=source_json,
            start_date=tournament["startDate"],
            players_count=tournament["attendeeCount"],
            points_limit=2000,
            rounds=5,
        )


def store_players(players, tournament_id):
    for player in players:
        source = SNL
        source_id = player["registrationID"]
        source_json = player

        player_instance, _ = Player.objects.update_or_create(
            source=source, source_id=source_id, source_json=source_json
        )

        participant, _ = Participant.objects.get_or_create(
            event=Event.objects.get(source_id=tournament_id),
            player=player_instance,
            source_json=player,
        )

        player_list, _ = List.objects.get_or_create(
            participant=participant,
            source_id=player["registrationID"],
            source_json=player,
            raw_list=str(player["listPlainText"]),
        )


def store_pairings(pairings, tournament_id):
    for pairing in pairings:
        source = SNL
        source_id = pairing["roundNumber"]
        source_json = pairing

        try:
            player1 = Player.objects.get(source=SNL, source_id=pairing["playerId1"])
            participant1 = Participant.objects.get(
                event__source_id=tournament_id, player=player1
            )
            player2 = Player.objects.get(source=SNL, source_id=pairing["playerId2"])
            participant2 = Participant.objects.get(
                event__source_id=tournament_id, player=player2
            )
        except Exception as e:
            print(
                f"Failed to store pairings for {tournament_id} error: {e} player1: {pairing['playerId1']} player2: {pairing['playerId2']}"
            )
            continue
        player1_list = List.objects.get(participant__player=player1)
        player2_list = List.objects.get(participant__player=player2)
        if pairing["playerResult1"] == "WIN":
            player1_result = "2"
            player2_result = "0"
        elif pairing["playerResult2"] == "WIN":
            player1_result = "0"
            player2_result = "2"
        else:
            player1_result = "1"
            player2_result = "1"

        if not Pairing.objects.filter(
            event__source_id=tournament_id,
            round=pairing["roundNumber"],
            source_id=source_id,
        ).exists():
            pairing = Pairing.objects.create(
                event=Event.objects.get(source_id=tournament_id),
                round=pairing["roundNumber"],
                source_id=source_id,
                source_json=pairing,
                player1=participant1,
                player2=participant2,
                player1_list=player1_list,
                player2_list=player2_list,
                player1_result=player1_result,
                player2_result=player2_result,
            )


def store_data(filename):
    tournaments, players, pairings = parse_snl_results_file(filename)

    # store_tournaments(tournaments)
    # for tournament_id, player_data in players.items():
    #     store_players(player_data, tournament_id)
    #
    for tournament_id, pairing_data in pairings.items():
        try:
            store_pairings(pairing_data, tournament_id)
        except Exception as e:
            print(f"Failed to store pairings for {tournament_id} error: {e}")
            continue


class Command(BaseCommand):
    help = "Fetch events from SNL"

    def handle(self, *args, **options):
        filename = "snl_results.xlsx"
        store_data(filename)
